import os
import openpyxl
import json

# ============ 内嵌模板数据（已清空示例商品） ============
embedded_template_data = [
    [
        {'value': '商品显示名称', 'font': {'name': '宋体', 'size': 11.0, 'bold': False, 'italic': False, 'color': None}, 'fill': {'fgColor': 'FFFFFF00', 'bgColor': None}, 'alignment': {'horizontal': 'center', 'vertical': 'center'}},
        {'value': '权重', 'font': {'name': '宋体', 'size': 11.0, 'bold': False, 'italic': False, 'color': None}, 'fill': {'fgColor': 'FFFFC000', 'bgColor': None}, 'alignment': {'horizontal': 'center', 'vertical': 'center'}},
        {'value': '货币显示名称', 'font': {'name': '宋体', 'size': 11.0, 'bold': False, 'italic': False, 'color': None}, 'fill': {'fgColor': 'FFFFFF00', 'bgColor': None}, 'alignment': {'horizontal': 'center', 'vertical': 'center'}},
        {'value': '货币实际名称', 'font': {'name': '宋体', 'size': 11.0, 'bold': False, 'italic': False, 'color': None}, 'fill': {'fgColor': 'FFFFC000', 'bgColor': None}, 'alignment': {'horizontal': 'center', 'vertical': 'center'}},
        {'value': '单价随机下限', 'font': {'name': '宋体', 'size': 11.0, 'bold': False, 'italic': False, 'color': None}, 'fill': {'fgColor': 'FFFFFF00', 'bgColor': None}, 'alignment': {'horizontal': 'center', 'vertical': 'center'}},
        {'value': '单价随机上限', 'font': {'name': '宋体', 'size': 11.0, 'bold': False, 'italic': False, 'color': None}, 'fill': {'fgColor': 'FFFFC000', 'bgColor': None}, 'alignment': {'horizontal': 'center', 'vertical': 'center'}},
        {'value': '个人限购', 'font': {'name': '宋体', 'size': 11.0, 'bold': False, 'italic': False, 'color': None}, 'fill': {'fgColor': 'FFFFFF00', 'bgColor': None}, 'alignment': {'horizontal': 'center', 'vertical': 'center'}},
        {'value': '全服限购', 'font': {'name': '宋体', 'size': 11.0, 'bold': False, 'italic': False, 'color': None}, 'fill': {'fgColor': 'FFFFC000', 'bgColor': None}, 'alignment': {'horizontal': 'center', 'vertical': 'center'}},
        {'value': '单次购买数量下限', 'font': {'name': '宋体', 'size': 11.0, 'bold': False, 'italic': False, 'color': None}, 'fill': {'fgColor': 'FFFFFF00', 'bgColor': None}, 'alignment': {'horizontal': 'center', 'vertical': 'center'}},
        {'value': '单次购买数量上限', 'font': {'name': '宋体', 'size': 11.0, 'bold': False, 'italic': False, 'color': None}, 'fill': {'fgColor': 'FFFFC000', 'bgColor': None}, 'alignment': {'horizontal': 'center', 'vertical': 'center'}},
        {'value': '限购实际数量', 'font': {'name': '宋体', 'size': 11.0, 'bold': False, 'italic': False, 'color': None}, 'fill': {'fgColor': 'FFFFFF00', 'bgColor': None}, 'alignment': {'horizontal': 'center', 'vertical': 'center'}},
        {'value': '是否自定义购买后的执行命令', 'font': {'name': '宋体', 'size': 11.0, 'bold': False, 'italic': False, 'color': None}, 'fill': {'fgColor': 'FFFFC000', 'bgColor': None}, 'alignment': {'horizontal': 'center', 'vertical': 'center'}},
        {'value': '商品实际名称', 'font': {'name': '宋体', 'size': 11.0, 'bold': False, 'italic': False, 'color': None}, 'fill': {'fgColor': 'FFFFFF00', 'bgColor': None}, 'alignment': {'horizontal': 'center', 'vertical': 'center'}},
        {'value': '商品数据值', 'font': {'name': '宋体', 'size': 11.0, 'bold': False, 'italic': False, 'color': None}, 'fill': {'fgColor': 'FFFFC000', 'bgColor': None}, 'alignment': {'horizontal': 'center', 'vertical': 'center'}},
        {'value': '自定义命令', 'font': {'name': '宋体', 'size': 11.0, 'bold': False, 'italic': False, 'color': None}, 'fill': {'fgColor': 'FFFFFF00', 'bgColor': None}, 'alignment': {'horizontal': 'center', 'vertical': 'center'}},
        {'value': '是否启用多次执行', 'font': {'name': '宋体', 'size': 11.0, 'bold': False, 'italic': False, 'color': None}, 'fill': {'fgColor': 'FFFFC000', 'bgColor': None}, 'alignment': {'horizontal': 'center', 'vertical': 'center'}},
    ],
    # 其他行保留样式但 value 为空
]

# 列宽和行高（保持和原文件一致）
col_widths = {
    'A': 20.0,
    'B': 15.0,
    'C': 18.0,
    'D': 18.0,
    'E': 15.0,
    'F': 15.0,
    'G': 15.0,
    'H': 15.0,
    'I': 20.0,
    'J': 20.0,
    'K': 15.0,
    'L': 25.0,
    'M': 18.0,
    'N': 15.0,
    'O': 25.0,
    'P': 20.0,
}
row_heights = {
    1: 25.0,
    2: 20.0,
}

def create_template_excel(path):
    wb = openpyxl.Workbook()
    ws = wb.active
    for col_letter, width in col_widths.items():
        if width:
            ws.column_dimensions[col_letter].width = width
    for row_idx, height in row_heights.items():
        if height:
            ws.row_dimensions[row_idx].height = height
    for r_idx, row in enumerate(embedded_template_data, start=1):
        for c_idx, cell in enumerate(row, start=1):
            cell_obj = ws.cell(row=r_idx, column=c_idx, value=cell["value"])
            if cell["font"]["name"]:
                cell_obj.font = openpyxl.styles.Font(
                    name=cell["font"]["name"],
                    size=cell["font"]["size"],
                    bold=cell["font"]["bold"],
                    italic=cell["font"]["italic"],
                    color=cell["font"]["color"]
                )
            if cell["fill"]["fgColor"]:
                cell_obj.fill = openpyxl.styles.PatternFill(
                    fill_type="solid",
                    fgColor=cell["fill"]["fgColor"]
                )
            if cell["alignment"]["horizontal"] or cell["alignment"]["vertical"]:
                cell_obj.alignment = openpyxl.styles.Alignment(
                    horizontal=cell["alignment"]["horizontal"],
                    vertical=cell["alignment"]["vertical"]
                )
    wb.save(path)
    print(f"已生成模板文件: {path}")

def excel_to_json_txt(path):
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    data = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(v is None for v in row):
            continue
        item = {headers[i]: (row[i] if row[i] is not None else "") for i in range(len(headers))}
        data.append(item)
    with open("商品清单.json", "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    print("已将 Excel 转为 商品清单.json")

if __name__ == "__main__":
    excel_path = "商品清单.xlsx"
    if not os.path.exists(excel_path):
        create_template_excel(excel_path)
    excel_to_json_txt(excel_path)
    input("\n按回车键退出程序...")
